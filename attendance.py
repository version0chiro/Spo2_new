import os
from datetime import date
import pandas as pd
from collections import Counter
import time
import xlsxwriter
def checkName(name,spo2,hr,Compensated,Ambient):



    def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):


        from openpyxl import load_workbook

        # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')

        writer = pd.ExcelWriter(filename, engine='openpyxl')

        # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
        try:
            FileNotFoundError
        except NameError:
            FileNotFoundError = IOError


        try:
            # try to open an existing workbook
            writer.book = load_workbook(filename)

            # get the last row in the existing Excel sheet
            # if it was not specified explicitly
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row

            # truncate sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = writer.book.sheetnames.index(sheet_name)
                # remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(sheet_name, idx)

            # copy existing sheets
            writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
        except FileNotFoundError:
            # file does not exist yet, we will create it
            pass

        if startrow is None:
            startrow = 0

        # write out the new sheet
        df.to_excel(writer, sheet_name,header=False, startrow=startrow, **to_excel_kwargs)

        # save the workbook
        writer.save()


    def add_to_list(str_name, atte, details,spo2,hr,Compensated,Ambient):
    	atte = atte
    	t = time.localtime()
    	current_time = time.strftime("%H:%M:%S", t)
    	d = details[details['Name'] == str_name]
    	d['Time-Stamp'] = current_time
    	d['SpO2_value'] = spo2
    	d['Heart-rate'] = hr
    	d['Body_temp'] = Compensated
    	d['Ambient'] = Ambient
        # d['SpO2_value']=spo2
    	# a = d.index[0]
    	return atte.append(d)

    def unknown_to_list(str_name, atte, details,spo2,hr,Compensated,Ambient):
    	atte = atte
    	t = time.localtime()
    	current_time = time.strftime("%H:%M:%S", t)
    	d = details[details['Name'] == str_name]
    	d['Name'] = str_name
    	d['Address'] = 'N/A'
    	d['Job'] = 'N/A'
    	d['Time-Stamp'] = current_time
    	d['SpO2_value'] = spo2
    	d['Heart-rate'] = hr
    	d['Body_temp'] = Compensated
    	d['Ambient'] = Ambient
        # d['SpO2_value']=spo2
    	a = d.index[0]
    	return atte.append(d)
 
    if os.path.isfile('excel_sheets/details.xlsx'): 
        details = pd.read_excel('excel_sheets/details.xlsx',engine="openpyxl",index_col=0)
    else:
        workbook = xlsxwriter.Workbook('excel_sheets/details.xlsx')
        workbook.close()
        details=pd.read_excel('excel_sheets/details.xlsx',engine="openpyxl",index_col=0)
    
    print(details)
    if os.path.isfile('excel_sheets/attendance.xlsx'): 
        atte=pd.read_excel('excel_sheets/attendance.xlsx',engine="openpyxl",index_col=0)
    else:
        workbook = xlsxwriter.Workbook('excel_sheets/attendance.xlsx')
        workbook.close()
        atte=pd.read_excel('excel_sheets/attendance.xlsx',engine="openpyxl",index_col=0)
        print(name)
    name = name.lower()
    print(name)
    if details['Name'].str.contains(name).any():
        print('testSuccess')
        try:
            atte = add_to_list(name,atte,details,spo2,hr,Compensated,Ambient)
            xl = pd.ExcelFile('excel_sheets/attendance.xlsx',engine="openpyxl")
            today = date.today()
            if str(today) in xl.sheet_names:
                append_df_to_excel('excel_sheets/attendance.xlsx',atte,sheet_name=str(today))
            else:
                with pd.ExcelWriter('excel_sheets/attendance.xlsx',engine="openpyxl",mode='a') as writer:
                    atte.to_excel(writer,sheet_name=str(today))
        except Exception as e:
            print(e)
            pass
    else:
        print('testSuccess')
        try:
            atte = unknown_to_list(name,atte,details,spo2,hr,Compensated,Ambient)
            xl = pd.ExcelFile('excel_sheets/attendance.xlsx',engine="openpyxl")
            today = date.today()
            if str(today) in xl.sheet_names:
                append_df_to_excel('excel_sheets/attendance.xlsx',atte,sheet_name=str(today))
            else:
                with pd.ExcelWriter('excel_sheets/attendance.xlsx',engine="openpyxl",mode='a') as writer:
                    atte.to_excel(writer,sheet_name=str(today))
        except Exception as e:
            print(e)
            pass
    
                            
